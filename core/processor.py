# -*- coding: utf-8 -*-
"""
엑셀 파일 처리 모듈 - 파일 읽기/쓰기 및 구조 분석
"""

import os
import pandas as pd
from pathlib import Path
from dataclasses import dataclass
from typing import List, Dict, Any, Optional, Tuple
import openpyxl
from openpyxl import load_workbook
import difflib
import config
import re

@dataclass
class SheetInfo:
    """시트 정보 데이터 클래스"""
    name: str
    header_row: int
    data_start_row: int
    columns: List[str]
    row_count: int
    data_types: Dict[str, str]
    sample_data: List[List[Any]]

@dataclass
class FileStructure:
    """파일 구조 정보 데이터 클래스"""
    file_path: str
    file_name: str
    sheets: List[SheetInfo]
    error_message: Optional[str] = None

class ExcelProcessor:
    """엑셀 파일 처리 클래스"""
    
    def __init__(self):
        """프로세서 초기화"""
        self.supported_formats = config.SUPPORTED_FORMATS
        
        # 동의어 사전 (한국어-한국어, 한국어-영어)
        self.synonym_dict = {
            # 이름 관련
            '이름': ['성명', 'name', '성함', '이름'],
            '성명': ['이름', 'name', '성함'],
            'name': ['이름', '성명', '성함'],
            
            # 나이 관련
            '나이': ['연령', 'age', '연세'],
            '연령': ['나이', 'age', '연세'],
            'age': ['나이', '연령', '연세'],
            
            # 부서 관련
            '부서': ['부서명', 'department', '소속', '소속부서', '팀'],
            '부서명': ['부서', 'department', '소속', '소속부서'],
            'department': ['부서', '부서명', '소속', '소속부서'],
            '소속부서': ['부서', '부서명', 'department'],
            
            # 날짜 관련
            '입사일': ['입사일자', 'hire_date', 'hiredate', '입사_일', '채용일'],
            '입사일자': ['입사일', 'hire_date', 'hiredate', '입사_일'],
            'hire_date': ['입사일', '입사일자', '채용일'],
            'hiredate': ['입사일', '입사일자', '채용일'],
            
            # 급여 관련
            '연봉': ['급여', 'salary', '월급', '연봉정보', '급료'],
            '급여': ['연봉', 'salary', '월급', '급료'],
            'salary': ['연봉', '급여', '월급', '급료'],
            '월급': ['연봉', '급여', 'salary', '급료'],
            
            # 일반적인 시트명
            '직원정보': ['employee_info', 'employee info', '사원정보', '직원 정보'],
            'employee_info': ['직원정보', '사원정보', '직원 정보'],
            'employee info': ['직원정보', '사원정보', '직원 정보'],
        }
        
    def read_file_structure(self, file_path: str) -> FileStructure:
        """
        엑셀/CSV 파일 구조 분석
        
        Args:
            file_path: 분석할 파일 경로
            
        Returns:
            FileStructure: 파일 구조 정보
        """
        try:
            # 파일 유효성 검사
            if not self._validate_file(file_path):
                return FileStructure(
                    file_path=file_path,
                    file_name=Path(file_path).name,
                    sheets=[],
                    error_message="유효하지 않은 파일입니다."
                )
            
            file_ext = Path(file_path).suffix.lower()
            sheets = []
            
            if file_ext == '.csv':
                # CSV 파일 처리
                sheet_info = self._analyze_csv_file(file_path)
                if sheet_info:
                    sheets.append(sheet_info)
            else:
                # 엑셀 파일 처리
                workbook = load_workbook(file_path, data_only=True)
                
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    sheet_info = self._analyze_sheet(worksheet, sheet_name)
                    if sheet_info:
                        sheets.append(sheet_info)
                
                workbook.close()
            
            return FileStructure(
                file_path=file_path,
                file_name=Path(file_path).name,
                sheets=sheets
            )
            
        except Exception as e:
            return FileStructure(
                file_path=file_path,
                file_name=Path(file_path).name,
                sheets=[],
                error_message=f"파일 읽기 오류: {str(e)}"
            )
    
    def _validate_file(self, file_path: str) -> bool:
        """
        파일 유효성 검사 (강화된 검증)
        
        Args:
            file_path: 검사할 파일 경로
            
        Returns:
            bool: 유효한 파일인지 여부
        """
        try:
            file_path_obj = Path(file_path)
            
            # 파일 존재 확인
            if not file_path_obj.exists():
                print(f"❌ 파일이 존재하지 않습니다: {file_path}")
                return False
            
            # 디렉토리 확인
            if file_path_obj.is_dir():
                print(f"❌ 디렉토리는 처리할 수 없습니다: {file_path}")
                return False
            
            # 파일 크기 확인
            try:
                file_size_mb = file_path_obj.stat().st_size / (1024 * 1024)
                if file_size_mb > config.MAX_FILE_SIZE_MB:
                    print(f"❌ 파일 크기가 너무 큽니다: {file_size_mb:.1f}MB (최대: {config.MAX_FILE_SIZE_MB}MB)")
                    return False
                elif file_size_mb == 0:
                    print(f"❌ 빈 파일입니다: {file_path}")
                    return False
            except Exception as e:
                print(f"❌ 파일 크기 확인 실패: {e}")
                return False
            
            # 확장자 확인
            file_ext = file_path_obj.suffix.lower()
            if file_ext not in config.SUPPORTED_FORMATS:
                print(f"❌ 지원하지 않는 파일 형식입니다: {file_ext} (지원 형식: {', '.join(config.SUPPORTED_FORMATS)})")
                return False
            
            # 파일 읽기 권한 확인
            try:
                with open(file_path, 'rb') as f:
                    f.read(1)  # 첫 바이트만 읽어서 권한 확인
            except PermissionError:
                print(f"❌ 파일 읽기 권한이 없습니다: {file_path}")
                return False
            except Exception as e:
                print(f"❌ 파일 접근 오류: {e}")
                return False
            
            # 파일 형식별 검증
            if file_ext == '.csv':
                return self._validate_csv_file(file_path_obj)
            else:
                return self._validate_excel_file(file_path_obj)
            
        except Exception as e:
            print(f"❌ 파일 유효성 검증 오류 ({file_path}): {e}")
            return False

    def _validate_excel_file(self, file_path: Path) -> bool:
        """엑셀 파일 특별 검증"""
        try:
            # 엑셀 파일 열기 테스트
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            
            # 시트 개수 확인
            if len(workbook.sheetnames) == 0:
                print(f"❌ 시트가 없는 엑셀 파일입니다: {file_path}")
                workbook.close()
                return False
            
            workbook.close()
            return True
            
        except Exception as e:
            print(f"❌ 엑셀 파일 검증 실패 ({file_path}): {e}")
            return False
    
    def _validate_csv_file(self, file_path: Path) -> bool:
        """CSV 파일 특별 검증"""
        try:
            import csv
            
            # 인코딩 테스트
            encodings = ['utf-8', 'cp949', 'euc-kr', 'latin-1']
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding, newline='') as f:
                        sample = f.read(1024)
                        if not sample.strip():
                            print(f"❌ 빈 CSV 파일입니다: {file_path}")
                            return False
                        
                        # CSV 형식 검증
                        f.seek(0)
                        sniffer = csv.Sniffer()
                        dialect = sniffer.sniff(sample)
                        
                        # 첫 번째 행 읽기 시도
                        f.seek(0)
                        reader = csv.reader(f, dialect=dialect)
                        first_row = next(reader, None)
                        
                        if not first_row or all(not str(cell).strip() for cell in first_row):
                            continue  # 다음 인코딩 시도
                        
                        return True
                        
                except UnicodeDecodeError:
                    continue
                except Exception:
                    continue
            
            print(f"❌ 읽을 수 없는 CSV 파일입니다: {file_path}")
            return False
            
        except Exception as e:
            print(f"❌ CSV 파일 검증 실패 ({file_path}): {e}")
            return False
    
    def _analyze_sheet(self, worksheet, sheet_name: str) -> Optional[SheetInfo]:
        """
        개별 시트 분석
        
        Args:
            worksheet: openpyxl 워크시트 객체
            sheet_name: 시트명
            
        Returns:
            SheetInfo: 시트 정보 (빈 시트인 경우 None)
        """
        try:
            # 시트 데이터 추출 (최대 100행까지만 분석)
            max_row = min(worksheet.max_row, 100) if worksheet.max_row else 0
            max_col = min(worksheet.max_column, 50) if worksheet.max_column else 0
            
            if max_row == 0 or max_col == 0:
                return None  # 빈 시트
            
            # 데이터 읽기 (빈 셀 처리 개선)
            data = []
            for row in worksheet.iter_rows(
                min_row=1, 
                max_row=max_row, 
                min_col=1, 
                max_col=max_col, 
                values_only=True
            ):
                # 빈 셀과 None 값을 명시적으로 처리
                processed_row = []
                for cell in row:
                    if cell is None or (isinstance(cell, str) and cell.strip() == ''):
                        processed_row.append(None)
                    else:
                        processed_row.append(cell)
                data.append(processed_row)
            
            # 헤더 행 탐지
            header_row = self._detect_header_row(data)
            
            if header_row == -1:
                return None  # 헤더를 찾을 수 없음
            
            # 열 정보 추출
            headers = data[header_row]
            columns = [str(h) if h is not None else f"열{i+1}" for i, h in enumerate(headers)]
            
            # 데이터 시작 행
            data_start_row = header_row + 1
            
            # 실제 데이터 행 수 계산
            actual_row_count = worksheet.max_row if worksheet.max_row else 0
            
            # 데이터 타입 분석
            data_types = self._analyze_data_types(data, header_row, columns)
            
            # 샘플 데이터 (최대 5행)
            sample_data = data[data_start_row:data_start_row + 5] if data_start_row < len(data) else []
            
            return SheetInfo(
                name=sheet_name,
                header_row=header_row,
                data_start_row=data_start_row,
                columns=columns,
                row_count=actual_row_count - data_start_row,
                data_types=data_types,
                sample_data=sample_data
            )
            
        except Exception as e:
            print(f"시트 분석 오류 ({sheet_name}): {e}")
            return None
    
    def _analyze_csv_file(self, file_path: str) -> Optional[SheetInfo]:
        """
        CSV 파일 분석
        
        Args:
            file_path: CSV 파일 경로
            
        Returns:
            SheetInfo: 시트 정보 (빈 파일인 경우 None)
        """
        try:
            import csv
            import io
            
            # 파일 인코딩 감지 및 데이터 읽기
            encodings = ['utf-8', 'cp949', 'euc-kr', 'latin-1']
            data = []
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding, newline='') as f:
                        # CSV 방언 감지
                        sample = f.read(1024)
                        f.seek(0)
                        
                        sniffer = csv.Sniffer()
                        try:
                            dialect = sniffer.sniff(sample)
                        except:
                            dialect = csv.excel  # 기본 방언 사용
                        
                        reader = csv.reader(f, dialect=dialect)
                        data = list(reader)
                        break
                except UnicodeDecodeError:
                    continue
            
            if not data:
                return None
            
            # 최대 100행까지만 분석
            data = data[:100]
            
            # 헤더 행 탐지
            header_row = self._detect_header_row(data)
            
            if header_row == -1:
                return None
            
            # 열 정보 추출
            headers = data[header_row] if header_row < len(data) else []
            columns = [str(h) if h else f"열{i+1}" for i, h in enumerate(headers)]
            
            # 데이터 시작 행
            data_start_row = header_row + 1
            
            # 실제 데이터 행 수 추정 (전체 파일 읽지 않고 추정)
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                row_count = sum(1 for line in f) - data_start_row
            
            # 데이터 타입 분석
            data_types = self._analyze_data_types(data, header_row, columns)
            
            # 샘플 데이터 (최대 5행)
            sample_data = data[data_start_row:data_start_row + 5] if data_start_row < len(data) else []
            
            return SheetInfo(
                name="Sheet1",  # CSV는 단일 시트로 처리
                header_row=header_row,
                data_start_row=data_start_row,
                columns=columns,
                row_count=max(row_count, 0),
                data_types=data_types,
                sample_data=sample_data
            )
            
        except Exception as e:
            print(f"CSV 파일 분석 오류: {e}")
            return None
    
    def _detect_header_row(self, data: List[List[Any]]) -> int:
        """
        헤더 행 자동 탐지
        
        각 행의 헤더 가능성 점수 계산:
        - 텍스트 비율 40%
        - 고유값 비율 30%
        - 빈 셀 비율 20% (역산)
        - 위치 점수 10% (상단 우선)
        
        Args:
            data: 시트 데이터
            
        Returns:
            int: 헤더 행 인덱스 (-1이면 헤더 없음)
        """
        if not data:
            return -1
        
        best_row = -1
        best_score = 0
        
        # 상위 10행만 검사
        for row_idx, row_data in enumerate(data[:10]):
            if not any(row_data):  # 빈 행 스킵
                continue
            
            # 유효한 셀 개수
            non_empty_cells = [cell for cell in row_data if cell is not None]
            if len(non_empty_cells) < 2:  # 너무 적은 데이터
                continue
            
            # 텍스트 비율
            text_count = sum(1 for cell in non_empty_cells if isinstance(cell, str))
            text_ratio = text_count / len(non_empty_cells)
            
            # 고유값 비율
            unique_values = set(non_empty_cells)
            unique_ratio = len(unique_values) / len(non_empty_cells)
            
            # 빈 셀 비율 (전체 행 기준)
            empty_count = sum(1 for cell in row_data if cell is None)
            empty_ratio = empty_count / len(row_data)
            
            # 위치 점수 (상단일수록 높음)
            position_score = (10 - row_idx) / 10
            
            # 총 점수 계산
            total_score = (
                text_ratio * 0.4 +
                unique_ratio * 0.3 +
                (1 - empty_ratio) * 0.2 +
                position_score * 0.1
            )
            
            if total_score > best_score:
                best_score = total_score
                best_row = row_idx
        
        # 최소 점수 임계값 확인
        return best_row if best_score > 0.5 else -1
    
    def _analyze_data_types(self, data: List[List[Any]], header_row: int, columns: List[str]) -> Dict[str, str]:
        """
        열별 데이터 타입 분석
        
        Args:
            data: 시트 데이터
            header_row: 헤더 행 인덱스
            columns: 열 이름 목록
            
        Returns:
            Dict[str, str]: 열별 데이터 타입 매핑
        """
        data_types = {}
        data_start = header_row + 1
        
        for col_idx, col_name in enumerate(columns):
            # 해당 열의 샘플 데이터 수집 (최대 10개)
            sample_values = []
            for row_idx in range(data_start, min(data_start + 10, len(data))):
                if col_idx < len(data[row_idx]) and data[row_idx][col_idx] is not None:
                    sample_values.append(data[row_idx][col_idx])
            
            if not sample_values:
                data_types[col_name] = 'text'
                continue
            
            # 데이터 타입 추론
            data_types[col_name] = self._infer_data_type(sample_values)
        
        return data_types
    
    def _infer_data_type(self, sample_values: List[Any]) -> str:
        """
        샘플 값들로부터 데이터 타입 추론
        
        Args:
            sample_values: 샘플 데이터 값들
            
        Returns:
            str: 추론된 데이터 타입 ('number', 'date', 'text')
        """
        import datetime
        
        number_count = 0
        date_count = 0
        text_count = 0
        
        for value in sample_values:
            if isinstance(value, (int, float)):
                number_count += 1
            elif isinstance(value, datetime.datetime):
                date_count += 1
            elif isinstance(value, str):
                # 문자열이지만 숫자로 변환 가능한지 확인
                try:
                    float(value.replace(',', ''))
                    number_count += 1
                except ValueError:
                    # 날짜 형식인지 확인
                    if self._is_date_string(value):
                        date_count += 1
                    else:
                        text_count += 1
            else:
                text_count += 1
        
        # 가장 많은 타입으로 결정
        total = len(sample_values)
        if number_count / total > 0.7:
            return 'number'
        elif date_count / total > 0.5:
            return 'date'
        else:
            return 'text'
    
    def _is_date_string(self, value: str) -> bool:
        """문자열이 날짜 형식인지 확인"""
        import re
        
        # 일반적인 날짜 패턴들
        date_patterns = [
            r'\d{4}-\d{1,2}-\d{1,2}',  # YYYY-MM-DD
            r'\d{4}/\d{1,2}/\d{1,2}',  # YYYY/MM/DD
            r'\d{1,2}/\d{1,2}/\d{4}',  # MM/DD/YYYY
            r'\d{1,2}-\d{1,2}-\d{4}',  # MM-DD-YYYY
            r'\d{4}\.\d{1,2}\.\d{1,2}', # YYYY.MM.DD
        ]
        
        for pattern in date_patterns:
            if re.match(pattern, value.strip()):
                return True
        
        return False
    
    def read_sheet_data(self, file_path: str, sheet_name: str, chunk_size: Optional[int] = None) -> pd.DataFrame:
        """
        특정 시트/CSV의 데이터를 DataFrame으로 읽기
        
        Args:
            file_path: 파일 경로
            sheet_name: 시트명 (CSV의 경우 무시됨)
            chunk_size: 청크 크기 (메모리 절약용)
            
        Returns:
            pd.DataFrame: 시트 데이터
        """
        try:
            file_ext = Path(file_path).suffix.lower()
            
            # CSV 파일 처리
            if file_ext == '.csv':
                return self._read_csv_data(file_path, chunk_size)
            
            # 엑셀 파일 처리
            structure = self.read_file_structure(file_path)
            
            if structure.error_message:
                raise Exception(structure.error_message)
            
            # 해당 시트 정보 찾기
            sheet_info = None
            for sheet in structure.sheets:
                if sheet.name == sheet_name:
                    sheet_info = sheet
                    break
            
            if not sheet_info:
                raise Exception(f"시트 '{sheet_name}'을 찾을 수 없습니다.")
            
            # pandas로 데이터 읽기 (빈 셀 안전 처리)
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=sheet_info.header_row,
                skiprows=None,
                nrows=chunk_size,
                keep_default_na=True,  # NaN 값 유지
                na_values=['', ' ', 'N/A', 'NA', 'NULL', 'null', '#N/A']  # 빈 값 처리
            )
            
            # 열 이름 정리 (빈 열 이름 처리)
            new_columns = []
            for i, col in enumerate(df.columns):
                if pd.isna(col) or str(col).strip() == '' or str(col).startswith('Unnamed:'):
                    new_columns.append(f'열{i+1}')
                else:
                    new_columns.append(str(col).strip())
            df.columns = new_columns
            
            return df
            
        except Exception as e:
            raise Exception(f"시트 데이터 읽기 오류: {str(e)}")
    
    def _read_csv_data(self, file_path: str, chunk_size: Optional[int] = None) -> pd.DataFrame:
        """
        CSV 파일을 DataFrame으로 읽기
        
        Args:
            file_path: CSV 파일 경로
            chunk_size: 청크 크기
            
        Returns:
            pd.DataFrame: CSV 데이터
        """
        try:
            import csv
            
            # 인코딩 자동 감지
            encodings = ['utf-8', 'cp949', 'euc-kr', 'latin-1']
            df = None
            
            for encoding in encodings:
                try:
                    # 구분자 자동 감지
                    with open(file_path, 'r', encoding=encoding, newline='') as f:
                        sample = f.read(1024)
                        sniffer = csv.Sniffer()
                        try:
                            delimiter = sniffer.sniff(sample).delimiter
                        except:
                            delimiter = ','  # 기본값
                    
                    # pandas로 읽기 (빈 셀 안전 처리)
                    df = pd.read_csv(
                        file_path,
                        encoding=encoding,
                        delimiter=delimiter,
                        nrows=chunk_size if chunk_size and chunk_size > 0 else None,
                        keep_default_na=True,
                        na_values=['', ' ', 'N/A', 'NA', 'NULL', 'null', '#N/A']
                    )
                    break
                    
                except UnicodeDecodeError:
                    continue
                except Exception as e:
                    continue
            
            if df is None:
                raise Exception("CSV 파일을 읽을 수 없습니다.")
            
            # 열 이름 정리 (빈 열 이름 처리)
            new_columns = []
            for i, col in enumerate(df.columns):
                if pd.isna(col) or str(col).strip() == '' or str(col).startswith('Unnamed:'):
                    new_columns.append(f'열{i+1}')
                else:
                    new_columns.append(str(col).strip())
            df.columns = new_columns
            
            return df
            
        except Exception as e:
            raise Exception(f"CSV 파일 읽기 오류: {str(e)}")
    
    def calculate_similarity(self, str1: str, str2: str) -> float:
        """
        두 문자열 간의 유사도 계산 (개선된 알고리즘 + 동의어 지원)
        
        Args:
            str1: 첫 번째 문자열
            str2: 두 번째 문자열
            
        Returns:
            float: 유사도 (0.0 ~ 1.0)
        """
        if not str1 or not str2:
            return 0.0
        
        # 문자열 정규화 (소문자, 공백/특수문자 제거)
        def normalize_string(s):
            s = str(s).lower().strip()
            # 공백, 하이픈, 언더스코어, 괄호 등 제거
            s = re.sub(r'[\s\-_\(\)\[\]\.]+', '', s)
            return s
        
        normalized_str1 = normalize_string(str1)
        normalized_str2 = normalize_string(str2)
        
        # 완전 일치 확인
        if normalized_str1 == normalized_str2:
            return 1.0
        
        # 동의어 검사
        synonym_score = self._check_synonyms(str1, str2)
        if synonym_score > 0:
            return synonym_score
        
        # 부분 포함 확인 (한쪽이 다른쪽을 포함)
        if normalized_str1 in normalized_str2 or normalized_str2 in normalized_str1:
            shorter_len = min(len(normalized_str1), len(normalized_str2))
            longer_len = max(len(normalized_str1), len(normalized_str2))
            return shorter_len / longer_len * 0.9  # 포함 관계일 때 높은 점수
        
        # difflib를 사용한 기본 유사도 계산
        basic_similarity = difflib.SequenceMatcher(None, normalized_str1, normalized_str2).ratio()
        
        # 추가 유사도 검사 - 공통 단어 비율
        words1 = set(re.findall(r'\w+', str1.lower()))
        words2 = set(re.findall(r'\w+', str2.lower()))
        
        if words1 and words2:
            common_words = words1.intersection(words2)
            word_similarity = len(common_words) / max(len(words1), len(words2))
            # 기본 유사도와 단어 유사도의 가중 평균
            return basic_similarity * 0.7 + word_similarity * 0.3
        
        return basic_similarity
    
    def _check_synonyms(self, str1: str, str2: str) -> float:
        """
        동의어 사전을 사용한 유사도 검사
        
        Args:
            str1: 첫 번째 문자열
            str2: 두 번째 문자열
            
        Returns:
            float: 동의어 유사도 (0.0 ~ 0.95)
        """
        str1_lower = str1.lower().strip()
        str2_lower = str2.lower().strip()
        
        # 직접 동의어 검사
        if str1_lower in self.synonym_dict:
            synonyms = [s.lower() for s in self.synonym_dict[str1_lower]]
            if str2_lower in synonyms:
                return 0.95  # 동의어는 높은 점수
        
        if str2_lower in self.synonym_dict:
            synonyms = [s.lower() for s in self.synonym_dict[str2_lower]]
            if str1_lower in synonyms:
                return 0.95  # 동의어는 높은 점수
        
        # 부분 동의어 검사 (단어 일부가 포함된 경우)
        for key, synonyms in self.synonym_dict.items():
            key_lower = key.lower()
            synonyms_lower = [s.lower() for s in synonyms]
            
            # str1이 key 또는 동의어 중 하나를 포함하고, str2도 마찬가지인 경우
            str1_matches = any(word in str1_lower for word in [key_lower] + synonyms_lower)
            str2_matches = any(word in str2_lower for word in [key_lower] + synonyms_lower)
            
            if str1_matches and str2_matches:
                return 0.8  # 부분 동의어는 중간 점수
        
        return 0.0
    
    def match_sheets(self, reference_sheets: List[str], target_sheets: List[str], threshold: float = None) -> Dict[str, str]:
        """
        시트명 매칭
        
        Args:
            reference_sheets: 기준 시트명 목록
            target_sheets: 대상 시트명 목록
            threshold: 유사도 임계값
            
        Returns:
            Dict[str, str]: 기준 시트명 -> 대상 시트명 매핑
        """
        if threshold is None:
            threshold = config.SIMILARITY_THRESHOLD
        
        matches = {}
        
        for ref_sheet in reference_sheets:
            best_match = None
            best_score = 0
            
            for target_sheet in target_sheets:
                score = self.calculate_similarity(ref_sheet, target_sheet)
                
                if score > best_score and score >= threshold:
                    best_score = score
                    best_match = target_sheet
            
            if best_match:
                matches[ref_sheet] = best_match
        
        return matches
    
    def match_columns(self, reference_columns: List[str], target_columns: List[str], threshold: float = None) -> Dict[str, str]:
        """
        열제목 매칭
        
        Args:
            reference_columns: 기준 열제목 목록
            target_columns: 대상 열제목 목록
            threshold: 유사도 임계값
            
        Returns:
            Dict[str, str]: 기준 열제목 -> 대상 열제목 매핑
        """
        if threshold is None:
            threshold = config.COLUMN_SIMILARITY_THRESHOLD
        
        matches = {}
        used_targets = set()
        
        for ref_col in reference_columns:
            best_match = None
            best_score = 0
            
            for target_col in target_columns:
                if target_col in used_targets:
                    continue
                
                score = self.calculate_similarity(ref_col, target_col)
                
                if score > best_score and score >= threshold:
                    best_score = score
                    best_match = target_col
            
            if best_match:
                matches[ref_col] = best_match
                used_targets.add(best_match)
        
        return matches 